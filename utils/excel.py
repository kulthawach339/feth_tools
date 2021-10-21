from openpyxl import load_workbook


class FileEntry:
    def __init__(self, file_id, strings):
        self.file_id = file_id
        self.strings = strings


def cell_is_num(value):
    try:
        int(value)
        return True
    except ValueError:
        return False
    except TypeError:
        return False


def parse_cell(cell):
    if cell is None:
        cell = ''
    return cell


def parse_excel(excel_path, is_debug=False):
    wb = load_workbook(excel_path)
    sh = wb.active
    entries = []
    for row_index in range(1, sh.max_row):
        raw = sh.cell(row=row_index, column=1).value
        translated = sh.cell(row=row_index, column=2).value

        if cell_is_num(translated):
            file_id = int(raw[:-4])
            lines_count = int(translated)
            strings = []
            for i in range(1, lines_count + 1):
                strings.append((parse_cell(sh.cell(row=row_index + i, column=1).value),
                                parse_cell(sh.cell(row=row_index + i, column=2).value),
                                parse_cell(sh.cell(row=row_index + i, column=3).value)))
            entry = FileEntry(file_id, strings)
            entries.append(entry)
            if is_debug:
                print(vars(entry))
    return entries
